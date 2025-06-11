import smtplib
import time
import random
import os
import csv
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
from dotenv import load_dotenv
from datetime import datetime

# Load environment variables
load_dotenv()
EMAIL = os.getenv("EMAIL")
PASSWORD = os.getenv("APP_PASSWORD")

# Load Excel sheet
workbook = load_workbook("Bytedance.xlsx")
sheet = workbook["Bytedance"]

# Extract headers and rows
headers = [cell.value for cell in sheet[1]]
required_columns = {"Name", "Email", "Company", "Role"}

# Validate Excel columns
if not required_columns.issubset(set(headers)):
    raise ValueError(f"❌ Missing required columns. Found: {headers}")

rows = list(sheet.iter_rows(min_row=2, values_only=True))
data = [dict(zip(headers, row)) for row in rows]

# Setup SMTP
# ✅ NEW: Gmail SMTP
server = smtplib.SMTP("smtp.gmail.com", 587)
server.starttls()
server.login(EMAIL, PASSWORD)


# Create or append to log file
LOG_FILE = "logs.csv"
if not os.path.exists(LOG_FILE):
    with open(LOG_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Email", "Status", "Message", "Timestamp"])

def log(email, status, msg):
    with open(LOG_FILE, "a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([email, status, msg, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])


def send_email(row):
    try:
        name = row["Name"].split()[0]
        company = row["Company"]
        role = row["Role"]
        email_to = row["Email"]
        link = row.get("Link", "")

        subject = f"Request for Interview - {role} at {company}"

        html = f"""
        <p>Hi {name},</p>
        <p>I’m <b>Abdul Rehman Ansari</b>, reaching out regarding the <b>{role}</b> role at <b>{company}</b>. Here's a quick summary of my qualifications:</p>
        <ul>
            <li><b>B.Tech in Mathematics and Computing</b> from Delhi Technological University (2021-2025)</li>
            <li><b>Technical Skills</b>: Python, C++, JavaScript, ReactJS, ExpressJS, SQL, and ML libraries (NumPy, Pandas)</li>
            <li><b>Project Experience</b>: Built AI Chrome extensions, ML models, and algorithmic solutions</li>
            <li><b>Achievements</b>: Fintestico Hackathon finalist, NPTEL Python Certification (86%)</li>
            <li>Available for immediate opportunities</li>
        </ul>
        <p>Links:</p>
        <ul>
            <li><a href="YOUR_RESUME_LINK">Resume</a></li>
            <li><a href="https://github.com/abrehman02">GitHub</a></li>
            {f'<li><a href="{link}">{role} Job Link</a></li>' if link else ""}
        </ul>
        <p>I’d love to discuss how I could contribute to your team. Looking forward to hearing from you.</p>
        <p>Best regards,<br>Abdul Rehman Ansari<br>📞 +91 9411942046</p>
        """

        msg = MIMEMultipart()
        msg["From"] = f"Abdul Rehman Ansari <{EMAIL}>"
        msg["To"] = email_to
        msg["Subject"] = subject
        msg.attach(MIMEText(html, "html"))

        server.send_message(msg)
        print(f"✅ Email sent to: {email_to}")
        log(email_to, "Success", "Email sent")


    except Exception as e:
        print(f"❌ Failed to send to {row.get('Email', 'N/A')}: {e}")
        log(row.get("Email", "N/A"), "Failed", str(e))

# Send emails with random delay
for row in data:
    send_email(row)
    wait = random.randint(30, 90)
    print(f"⏱️ Waiting {wait}s before next email...")
    time.sleep(wait)

server.quit()
print("🎉 All emails processed.")